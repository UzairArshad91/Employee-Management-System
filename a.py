import openpyxl
from openpyxl import workbook
import customtkinter as ctk
from customtkinter import *
import tkinter as tk
from tkinter import simpledialog, messagebox
from tkinter import ttk
import os

#openpxyl excel file sync

excelFile='employees.xlsx'

wb=openpyxl.load_workbook(excelFile)

ws=wb['Employees']


#functions 
#need 4 FOR BUTTONS

#func 1 find the row for employee

def findEmployee(employeeID):
    for row in ws.iter_rows(min_row=2, values_only=False):
         #row is a variable that will iterate in rows of worksheet starting from row 2 since roe 1 is header
        
        if str(row[0].value)==str(employeeID): 
            #if input ID matches row[0] it will return the row number. row[0] bcs 1st index is ID of employees
            return row
        
    return None

#func 2 saves the workbook
def workbookSave():

    wb.save(excelFile) 
    #openpyxl func to save workbook

#func 3 for button: display employees list
def displayEmployee():
    top=ctk.CTkToplevel(root) #top=new window on top of root window
    top.title("Employees list")
    top.geometry("800x600")

    top.transient(root)   #Makes window stay on top of main root  

    style = ttk.Style()
    style.configure("Treeview", font=("Arial", 12))          # Row font
    style.configure("Treeview.Heading", font=("Arial", 12, "bold"))  # Header font

 
    tree=ttk.Treeview(top, columns=("id","Name","Post","salary"), show="headings") 
    #show headings mean that heading below and their name will be shown actually

    
    #text will be shown, first is id like html
    tree.heading("id",text="id")
    tree.heading("Name",text="Names")
    tree.heading("Post", text="Post")
    tree.heading("salary",text="Salary")

    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert("","end",values=row) 
        #skipping row 1 bcs we have already printed headers
        #insert: insert a new row every iteration, "":not nested rows, end: add to bottom, displays worksheet values in row

    tree.pack(fill="both", expand=True)

    #both: will expand in x and y axis

#func 4 for button: add an employee 
#access workbook and check if employee already exists or no. If not then take further info for employee and save it in workbook
def addEmployee():
    lastId = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            try:
                lastId = int(row[0]) #explicit type casting from string to integer
                #doesn;t work without try except
            except ValueError:
                lastId = 0  

    newId = lastId + 1


    Name=simpledialog.askstring("Add Employee:", "Enter Employee Name:", parent=root)
    Post=simpledialog.askstring("Add Employee:", "Enter Employee Post:", parent=root)
    
    if not Name or not Post:
        messagebox.showerror("Error", "Name and Post are required.", parent=root)
        return
    try:
        salary=float(simpledialog.askstring("Add employee","Enter Employee salary", parent=root))

    except:
        messagebox.showerror("Error","Invalid Salary", parent=root)
        return
    
    ws.append([newId,Name,Post,salary])
    workbookSave()

    messagebox.showinfo("Success","Employee has been added", parent=root)


#func 5 for button: Promote an employee
#func 5 for button: Promote an employee
def promoteEmployee():
    id=simpledialog.askstring("Promote an employee","Enter Employee ID:", parent=root)
    if not id: return
    #to check if employee already exists or not

    row=findEmployee(id)
    if not row:
        messagebox.showerror("Error","Employee not found", parent=root)
        return
    
    try:
        amount=int(simpledialog.askstring("Promote an employee","Enter salary increment amount", parent=root))
        post=simpledialog.askstring("Promote an employee","Enter new Post", parent=root)
    except:
        messagebox.showerror("Error","Invalid input", parent=root)
        return
    
    currentSalary=int(row[3].value)
    row[3].value=currentSalary+amount

    #Keep old post if user leaves input empty or presses cancel
    if post and post.strip() != "":
        row[2].value = post

    workbookSave()
    messagebox.showinfo("Success!","Employee promoted!", parent=root)


#func 6 for button: remove an employee
def removeEmployee():
    id=simpledialog.askstring("Remove an employee","Enter employee ID to remove", parent=root)
    if not id: return
    #checking if employee exists

    DelRow=findEmployee(id)
    #this will save the row number of employee to be deleted in DelRow
    if not DelRow:
        messagebox.showerror("Error","Employee not found", parent=root)
        return
    
    confirm = messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete Employee ID {id}?", parent=root)
    if not confirm:
        return

    ws.delete_rows(DelRow[0].row) 
    #openyxl func to delete row
    workbookSave()
    messagebox.showinfo("Success","Employee successfully deleted", parent=root)

def showAbout():
    messagebox.showinfo("About", "Employee Management System\nVersion 1.0\nBy M Uzair Bin Arshad", parent=root)


#gui window
root = ctk.CTk()
root.title("Employee management system")
root.geometry("500x450")
set_appearance_mode("dark") 

titleLabel=ctk.CTkLabel(root, text="Employee management system", text_color="#fc001d", font=ctk.CTkFont(size=20, weight="bold")).pack(pady=30)

ctk.CTkButton(root, text="Display Employees",corner_radius=50 ,text_color="#FFFFFF", hover_color="#475c1e",width=200, command=displayEmployee,).pack(pady=10)

ctk.CTkButton(root, text="Add Employee",corner_radius=50 ,text_color="#FFFFFF", hover_color="#475c1e", width=200, command=addEmployee).pack(pady=10)

ctk.CTkButton(root, text="Promote Employee",corner_radius=50 ,text_color="#FFFFFF", hover_color="#475c1e", width=200,  command=promoteEmployee).pack(pady=10)

ctk.CTkButton(root, text="Remove Employee",corner_radius=50 ,text_color="#FFFFFF", hover_color="#475c1e", width=200, command=removeEmployee).pack(pady=10)

ctk.CTkButton(root, text="Exit Program",corner_radius=50 ,text_color="#FFFFFF", hover_color="#8a0111", width=200, command=root.quit).pack(pady=20)

ctk.CTkButton(root, text="About", corner_radius=50, text_color="#FFFFFF", hover_color="#475c1e", width=200, command=showAbout).pack(pady=10)

root.eval('tk::PlaceWindow . center')
root.resizable(False, False)

root.mainloop()

